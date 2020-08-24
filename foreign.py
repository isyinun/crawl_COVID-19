import requests
import json
import time
import openpyxl

#Get_foreign=r"https://view.inews.qq.com/g2/getOnsInfo?name=disease_other"
Get_foreign=r"https://view.inews.qq.com/g2/getOnsInfo?name=disease_foreign"
Get_China=r"https://view.inews.qq.com/g2/getOnsInfo?name=disease_h5"
Get_foreign_all=r"https://api.inews.qq.com/newsqa/v1/automation/foreign/country/ranklist"
class item:
    def __init__(self):
        self.country=list()#国家
        self.province = list()#省份
        self.area=list()#地区
        self.confirm=list()#确诊
        self.suspect=list()#疑似
        self.heal=list()#治愈
        self.dead=list()#死亡
        self.add=list()#新增
Data_Box=item()#数据盒子

def GetHtmlText(url):
    try:
        res = requests.get(url,timeout = 30)
        res.raise_for_status()
        res.encoding = res.apparent_encoding
        return res.text
    except:
        return "Error"
Foreign = GetHtmlText(Get_foreign)
Fore_Count_json = json.loads(Foreign)
Fore_Count_json = Fore_Count_json["data"] #解析data字段数据
Fore_Count_json = json.loads(Fore_Count_json) #转换为json字符串

Foreign_all = GetHtmlText(Get_foreign_all)
Foreall_Count_json = json.loads(Foreign_all)
Foreall_Count_json = Foreall_Count_json["data"] #解析data字段数据
#Foreall_Count_json = json.loads(Foreall_Count_json) #转换为json字符串
print(Foreall_Count_json)   

#China = GetHtmlText(Get_China)
#City_Count_json = json.loads(China)
#City_Count_json = City_Count_json["data"]#将json数据中的data字段的数据提取处理
#City_Count_json = json.loads(City_Count_json)#将提取出的字符串转换为json数据
#lastUpdateTime = City_Count_json["lastUpdateTime"]

#foreignList_json = Fore_Count_json["foreignList"]  #国外疫情信息包括国家地区
foreignList_json = Foreall_Count_json #各个国家分别统计
globalStatis_json = Fore_Count_json["globalStatis"]
global_nowtotal = str(globalStatis_json["nowConfirm"]) #国外现有确诊
global_total = str(globalStatis_json["confirm"]) #国外累计确诊
global_heal = str(globalStatis_json["heal"]) #国外累计治愈
global_dead = str(globalStatis_json["dead"]) #国外累计死亡
lastUpdateTime = str(globalStatis_json["lastUpdateTime"])
print(globalStatis_json)
#print(foreignList_json)
#foreign_len = len(foreignList_json)
#print(foreign_len)
def Get_Data_Foreign():
    foreign_len = len(foreignList_json)
    print(foreign_len)
    for i in range(0,foreign_len):
        f_name = foreignList_json[i]["name"]
        f_confirm = foreignList_json[i]["confirm"]
        f_add = foreignList_json[i]["confirmAdd"]
        f_dead = foreignList_json[i]["dead"]
        f_heal = foreignList_json[i]["heal"]
        Data_Box.country.append(f_name)
        Data_Box.province.append(f_name)
        Data_Box.area.append(f_name)
        Data_Box.confirm.append(f_confirm)
        Data_Box.heal.append(f_heal)
        Data_Box.dead.append(f_dead)
        Data_Box.add.append(f_add)
        #print(Data_Box.confirm)
    return len(Data_Box.country)
f_length = Get_Data_Foreign()
#print(f_length)
def write(length):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, value="国家")
    ws.cell(1, 2, value="省份")
    ws.cell(1, 3, value="地区")
    ws.cell(1, 4, value="确诊人数")
    ws.cell(1, 5, value="治愈人数")
    ws.cell(1, 6, value="死亡人数")
    ws.cell(1, 7, value="新增确诊")
    for n in range(0,length):
        ws.cell(n + 2, 1, Data_Box.country[n])
        ws.cell(n + 2, 2, Data_Box.province[n])
        ws.cell(n + 2, 3, Data_Box.area[n])
        ws.cell(n + 2, 4, Data_Box.confirm[n])
        ws.cell(n + 2, 5, Data_Box.heal[n])
        ws.cell(n + 2, 6, Data_Box.dead[n])
        ws.cell(n + 2, 7, Data_Box.add[n])
    ws.cell(n + 3, 1, value="海外现有确诊人数：" + global_nowtotal)
    ws.cell(n + 4, 1, value="海外累计确诊人数：" + global_total)
    ws.cell(n + 5, 1, value="海外累计治愈人数：" + global_heal)
    ws.cell(n + 6, 1, value="海外累计死亡人数：" + global_dead)
    ws.cell(n + 7, 1, value="更新时间：" + lastUpdateTime)
    name = 'f'+lastUpdateTime + '.xlsx'
    #name = name.replace(':', '_')
    #wb.save("D:/data2019-nCoV/{}".format(name))
    wb.save("/Users/apple/Downloads/SARS/{}".format(name))
write(f_length)
